"""
Excel Generator - Renteberekening Export
=========================================
Genereert een .xlsx bestand met 3 tabbladen:
1. Samenvatting - Overzicht vorderingen, status, openstaand
2. Periodes - Elke renteperiode per vordering
3. Betalingen - Toerekening per betaling
"""

from io import BytesIO
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


RENTETYPE_LABELS = {
    1: 'Wettelijk samengesteld',
    2: 'Handelsrente samengesteld',
    3: 'Wettelijk enkelvoudig',
    4: 'Handelsrente enkelvoudig',
    5: 'Contractueel vast',
    6: 'Wettelijk + opslag',
    7: 'Handelsrente + opslag',
}

HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(name='Calibri', bold=True, color="ffffff", size=10)
DATA_FONT = Font(name='Calibri', size=10)
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.00%'
DATE_FORMAT = 'DD-MM-YYYY'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='d0d0d0'),
)
TOTAL_FILL = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
TOTAL_FONT = Font(name='Calibri', bold=True, size=10)


def _style_header_row(ws, num_cols: int):
    """Style the header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 3, 30)
        ws.column_dimensions[col[0].column_letter].width = adjusted


def generate_excel(invoer: Dict[str, Any], resultaat: Dict[str, Any]) -> bytes:
    """Generate Excel workbook with 3 sheets."""
    wb = Workbook()

    _build_samenvatting(wb, invoer, resultaat)
    _build_periodes(wb, resultaat)
    _build_betalingen(wb, resultaat)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _build_samenvatting(wb: Workbook, invoer: Dict, resultaat: Dict):
    """Build Samenvatting sheet."""
    ws = wb.active
    ws.title = "Samenvatting"

    # Case info
    case = invoer.get('case', {})
    ws.append(["Renteberekening", None, None, None])
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws.append(["Einddatum:", case.get('einddatum', '')])
    ws.append(["Strategie:", 'Meest bezwarend (art. 6:43 BW)' if case.get('strategie') == 'A' else 'Oudste eerst'])
    ws.append([])

    # Vorderingen overzicht
    headers = ["Kenmerk", "Type", "Bedrag", "Kosten", "Rente", "Rente kosten", "Afgelost", "Openstaand", "Status"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    vorderingen = resultaat.get('vorderingen', [])
    for v in vorderingen:
        item_label = "(K) " if v.get('item_type') == 'kosten' else ""
        row = [
            item_label + v.get('kenmerk', ''),
            RENTETYPE_LABELS.get(v.get('rentetype', 1), '?'),
            float(v.get('oorspronkelijk_bedrag', 0)),
            float(v.get('kosten', 0)),
            float(v.get('totale_rente', 0)),
            float(v.get('totale_rente_kosten', 0)),
            float(v.get('afgelost_hoofdsom', 0)) + float(v.get('afgelost_kosten', 0)) + float(v.get('afgelost_rente', 0)) + float(v.get('afgelost_rente_kosten', 0)),
            float(v.get('openstaand', 0)),
            v.get('status', 'OPEN'),
        ]
        ws.append(row)
        r = ws.max_row
        for col in [3, 4, 5, 6, 7, 8]:
            ws.cell(row=r, column=col).number_format = MONEY_FORMAT
        ws.cell(row=r, column=col).border = THIN_BORDER

    # Totalen
    totalen = resultaat.get('totalen', {})
    ws.append([])
    total_row = [
        "TOTAAL",
        "",
        float(totalen.get('oorspronkelijk', 0)),
        float(totalen.get('kosten', 0)),
        float(totalen.get('rente', 0)),
        float(totalen.get('rente_kosten', 0)),
        float(totalen.get('afgelost_hoofdsom', 0)) + float(totalen.get('afgelost_kosten', 0)) + float(totalen.get('afgelost_rente', 0)) + float(totalen.get('afgelost_rente_kosten', 0)),
        float(totalen.get('openstaand', 0)),
        "",
    ]
    ws.append(total_row)
    r = ws.max_row
    for col in range(1, len(total_row) + 1):
        ws.cell(row=r, column=col).fill = TOTAL_FILL
        ws.cell(row=r, column=col).font = TOTAL_FONT
    for col in [3, 4, 5, 6, 7, 8]:
        ws.cell(row=r, column=col).number_format = MONEY_FORMAT

    _auto_width(ws)


def _build_periodes(wb: Workbook, resultaat: Dict):
    """Build Periodes sheet - one row per interest period per vordering."""
    ws = wb.create_sheet("Periodes")

    headers = ["Vordering", "Van", "Tot", "Dagen", "Hoofdsom/Kosten", "Rente %", "Rente", "Kapitalisatie", "Pauze"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for v in resultaat.get('vorderingen', []):
        kenmerk = v.get('kenmerk', '')

        # Hoofdsom periodes
        for p in v.get('periodes', []):
            row = [
                kenmerk,
                p.get('start', ''),
                p.get('eind', ''),
                f"{p.get('dagen', 0)}/{p.get('dagen_jaar', 365)}",
                float(p.get('hoofdsom', 0)),
                float(p.get('rente_pct', 0)),
                float(p.get('rente', 0)),
                "Ja" if p.get('is_kapitalisatie') else "",
                "Ja" if p.get('is_pauze') else "",
            ]
            ws.append(row)
            r = ws.max_row
            ws.cell(row=r, column=5).number_format = MONEY_FORMAT
            ws.cell(row=r, column=6).number_format = PCT_FORMAT
            ws.cell(row=r, column=7).number_format = MONEY_FORMAT

        # Kosten periodes
        for p in v.get('periodes_kosten', []):
            row = [
                f"{kenmerk} (kosten)",
                p.get('start', ''),
                p.get('eind', ''),
                f"{p.get('dagen', 0)}/{p.get('dagen_jaar', 365)}",
                float(p.get('kosten', 0)),
                float(p.get('rente_pct', 0)),
                float(p.get('rente', 0)),
                "",
                "Ja" if p.get('is_pauze') else "",
            ]
            ws.append(row)
            r = ws.max_row
            ws.cell(row=r, column=5).number_format = MONEY_FORMAT
            ws.cell(row=r, column=6).number_format = PCT_FORMAT
            ws.cell(row=r, column=7).number_format = MONEY_FORMAT

    _auto_width(ws)


def _build_betalingen(wb: Workbook, resultaat: Dict):
    """Build Betalingen sheet - payment allocations."""
    ws = wb.create_sheet("Betalingen")

    headers = ["Betaling", "Datum", "Bedrag", "Verwerkt", "Vordering", "Type", "Toegerekend"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for d in resultaat.get('deelbetalingen', []):
        kenmerk = d.get('kenmerk', '-')
        datum = d.get('datum', '')
        bedrag = float(d.get('bedrag', 0))
        verwerkt = float(d.get('verwerkt', 0))

        toerekeningen = d.get('toerekeningen', [])
        if not toerekeningen:
            ws.append([kenmerk, datum, bedrag, verwerkt, "", "", 0])
            r = ws.max_row
            for col in [3, 4, 7]:
                ws.cell(row=r, column=col).number_format = MONEY_FORMAT
        else:
            for i, t in enumerate(toerekeningen):
                type_labels = {
                    'hoofdsom': 'Hoofdsom',
                    'rente': 'Rente',
                    'kosten': 'Kosten',
                    'rente_kosten': 'Rente op kosten',
                }
                row = [
                    kenmerk if i == 0 else "",
                    datum if i == 0 else "",
                    bedrag if i == 0 else "",
                    verwerkt if i == 0 else "",
                    t.get('vordering', ''),
                    type_labels.get(t.get('type', ''), t.get('type', '')),
                    float(t.get('bedrag', 0)),
                ]
                ws.append(row)
                r = ws.max_row
                for col in [3, 4, 7]:
                    ws.cell(row=r, column=col).number_format = MONEY_FORMAT

    _auto_width(ws)
